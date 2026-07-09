# [네이버 뉴스 검색 Open API ] 에서 AI 관련 기사 10개를 가져와 엑셀파일로 자동 저장하기
# [작업 흐름 Workflow]
# 1. 네이버 개발자 사이트 [https://developers.naver.com] 에 로그인.
# 2. 내 애플리케이션 등록 [서비스 환경추가 : web- 도메일 ‘http://localhost’ 또는 호스팅 웹서버]
# 3. 애플리케이션 등록하면 발급되는 인증키 CLIENT_ID, CLIENT_SECRET 확인
# 4. 검색 Open API 사용을 위한 가이드 문서 참고하여 [뉴스] 데이터 10개를 json 형식으로 요청
# 5. 뉴스의 데이터들을 모두 엑셀에 저장.
# 6. 엑셀의 제목줄은 ["no", "title", "originallink", "link", "description", "pubDate"] 형태로 저장
# 7. 엑셀파일로 저장 - 매일 1번 저장하는 내용이기에 저장되는 날짜로 파일명 설정.
# [ 저장경로 : /news_archive/ , 파일명 : 20260709.xlsx ]
# ※ 실제로 매일 자동으로 저장되는 시스템까지는 구축하지 않음.
# ※ 하루에 한번 앱을 직접 실행하여 매일 뉴스가 저장되도록 한다는 가정으로 구현.

# import requests

# # naver new API 
# url = 'https://openapi.naver.com/v1/search/news.json'
# params={
#     'query':''
# }
# response = requests.get(url,params=params) # GET 방식 요청
# print(response.text)


# 네이버 검색 API 예제 - 블로그 검색
import os
import sys
import urllib.request
client_id = 'nsACDaGp9iUA7WFmbPhm' # "YOUR_CLIENT_ID"
client_secret = 'ffmHdo2PA9'  #"YOUR_CLIENT_SECRET"
encText = urllib.parse.quote("검색할 단어")
url = "https://openapi.naver.com/v1/search/blog?query=" + encText # JSON 결과
# url = "https://openapi.naver.com/v1/search/blog.xml?query=" + encText # XML 결과
request = urllib.request.Request(url)
request.add_header("X-Naver-Client-Id",client_id)
request.add_header("X-Naver-Client-Secret",client_secret)
response = urllib.request.urlopen(request)
rescode = response.getcode()
if(rescode==200):
    response_body = response.read()
    # print(response_body.decode('utf-8'))    
else:
    print("Error Code:" + rescode)






from openpyxl import Workbook
import time
wb = Workbook()
search_time =time.localtime
sheet = wb.create_sheet(str(search_time))

sheet.append(["no", "title", "originallink", "link", "description", "pubDate"])
import json
json_data = json.loads(response_body.decode('utf-8'))
# print(json_data)
# print(json_data['items'])
items = json_data['items']
print(len(items))
# print(json_data)

count=0
for idx in items[0]:    
    print(idx,end='\t\t')
count +=1




count =1
for lines in json_data['items']:
    #  sheet.append(lines)    
    col=1
    for idx in lines:
        
        print(idx)
        print(lines[idx])
        if col==1:
            sheet.cell(row=count+1,column=col).value=count  
        else:            
            if idx=='title' :
                title = lines["title"].replace("<b>", "").replace("</b>", "")
                title = lines["title"].replace("&quot;", "").replace("&quot;", "")
                lines[idx]=title
            if idx == 'description':
                description = lines["description"].replace("<b>", "").replace("</b>", "")
                # description = lines["description"].replace("&quot;", "").replace("&quot;", "")
                lines[idx]=description
            sheet.cell(row=count+1,column=col).value=lines[idx]        
        col+=1
    count +=1



wb.save('./news_archive/20260709.xlsx')
wb.close()