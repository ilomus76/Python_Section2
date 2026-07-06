from openpyxl import Workbook
from openpyxl import load_workbook

# wb = Workbook()
wb=load_workbook('./scores.xlsx')
sheet = wb.active

#퀴즈2 라는 제목을 찾자
# for row_num in range(1,12):
#     for col_num in range(1,10):
#         if sheet.cell(row=row_num,col=col_num).value=='퀴즈2':
#            cell_row=row_num
#            cell_col=col_num
           
#===============================================
target_col=0
for n in range(1,10):
    print(sheet.cell(1,n).value)
    if sheet.cell(1,n).value =="퀴즈2":
        target_col = n 

for i in range(2,12):
    sheet.cell(i,target_col).value=10
#================================================

for n in range(1,10):
    if sheet.cell(1,n).value == '총점':
        target_col= n


total =0
avg=0

for i in range(2,12):
    for k in range(3,target_col):
        total += int(sheet.cell(i,k).value);    
        
        
    sheet.cell(i,target_col).value= total
    # avg = sheet.cell(i,target_col).value/5
    # avg = total/5
    # avg=70

    if total >= 90:
         sheet.cell(i,target_col+1).value='A'
    elif 80<=total<90:
         sheet.cell(i,target_col+1).value='B'
    elif 70<=total<80:
         sheet.cell(i,target_col+1).value='C'
    else:
         sheet.cell(i,target_col+1).value='D'
    total=0
    avg =0 

#==================================================

for i in range(1,10):
    if sheet.cell(1,i).value=='출석':
        target_col = i

print(target_col)

for i in range(2,12):
    if sheet.cell(i,target_col).value <5 : 
        sheet.cell(i,9).value = 'F'





# 1. 퀴즈2 점수를 10 으로 수정
# 2. H열에 총점(SUM 이용), I열에 성적 정보 추가
#   - 총점 90 이상 A, 80 이상 B, 70 이상 C, 나머지 D
# 3. 출석이 5 미만인 학생은 총점 상관없이 F





# print(sheet[1].value)


# for n in sheet[1]:
#     print(n)

# print(sheet.cell(row=1,column=1).value) #'A1'
# print(sheet.cell(row=1,column=2).value) #'B1'



# print(sheet.rows[0])


# for topic_cols in sheet.rows[0]:
#     for idx in topic_cols:
#         if sheet[idx] == "퀴즈2":
#             # sheet.iter_cols(1,10) = 10\\\
#             # sheet.(min_rows=2, max_rows)
#             print(sheet[idx])
#         print()
#     print()
# print()

# # for row in sheet.iter_cols(min_row=2): # 제목줄 제외 
# for row in sheet.iter_rows(min_row=2): # 제목줄 제외 
#     if int(row[1].value)>80: #국어 성적이 80점 이상인 경우 
#         print(row[0].value,"번 학생은 국어 성적이 매우 우수합니다.")

# #'영어'과목명이 써있는 cell을 찾아 'english'로 변경하여 'sample4_modify.xlsx'로 저장

# row = sheet[1] #제목줄
# for cell in row:
#     if cell.value =='영어':
#         cell.value = 'english'

# wb.save('./01_excel/sample4_modify.xlsx')
# wb.close()






wb.save('./score_modify.xlsx')
wb.close()


