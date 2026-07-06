#병합된 셀 분리

from openpyxl import load_workbook

wb = load_workbook('./01_excel/sample6.xlsx')

sheet = wb.active

sheet.unmerge_cells('B2:D2')

wb.save('./01_excel/sample6_unmerge.xlsx')
wb.close()