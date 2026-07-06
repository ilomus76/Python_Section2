from openpyxl import Workbook

wb=Workbook()

sheet =wb.active
sheet.title = 'RPA sheet'

#1. 특정 cell에 값 입력
sheet['A1']=10
sheet['A2']=20
sheet['A3']=30

sheet['B1']=40
sheet['B2']=50
sheet['B3']=60

#2.특정 cell의 값 가져와서 출력해보기
print(sheet['A1'].value)
print(sheet['A5'].value) # 값이 없는 셀의 value : None : JS에서는 undefined

#3. 셀 접근할때.. '셀이름' eotls 행열 위치번호(숫자)로 접근하기 .. python과 다르게 1부터... 데이타베이스 및 엑셀같은 것은 1부터 시작
#row : 1,2,3
#column : 1(A),2(B),3(c),......

print(sheet.cell(row=1,column=1).value) #'A1'
print(sheet.cell(row=1,column=2).value) #'B1'

#4. 특정 셀의 위치에 값 지정하는 것도 위치번호로 가능

sheet.cell(row=1,column=3,value=70) #C1 셀에 70을 입력

#5. 셀 번호를 이용할때 장점. 반복문 값을 처리하기 용이.
import random

for n in range(1,11): #1~10
    num = random.randint(0,100) # 0~100까지의 랜덤한 숫자
    sheet.cell(row=n,column=4,value=num)




# 저장
wb.save('./01_excel/sample3.xlsx')
wb.close()