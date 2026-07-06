#특정 셀 찾기 
from openpyxl import load_workbook

wb = load_workbook('./01_excel/sample4.xlsx')
sheet = wb.active

#각 줄마다 접근하여 80점 이상의 학생 데이타 찾기
# for row in sheet.iter_cols(min_row=2): # 제목줄 제외 
for row in sheet.iter_rows(min_row=2): # 제목줄 제외 
    if int(row[1].value)>80: #국어 성적이 80점 이상인 경우 
        print(row[0].value,"번 학생은 국어 성적이 매우 우수합니다.")

#'영어'과목명이 써있는 cell을 찾아 'english'로 변경하여 'sample4_modify.xlsx'로 저장

row = sheet[1] #제목줄
for cell in row:
    if cell.value =='영어':
        cell.value = 'english'

wb.save('./01_excel/sample4_modify.xlsx')
wb.close()
