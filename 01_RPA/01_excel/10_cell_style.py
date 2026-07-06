from openpyxl import load_workbook
wb = load_workbook('./01_excel/sample4.xlsx')
sheet = wb.active

#1. A열의 너비를 5로 설정. 
sheet.column_dimensions['A'].width=5

#2. 행의 높이 지정
sheet.row_dimensions[1].height=50

#3.셀 스타일 [글씨색상, 셀 배경색] -- 하위모듈에 있음
from openpyxl.styles import Font
sheet['A1'].font= Font(color='FF0000', italic=True, bold=True)
sheet['B1'].font= Font(color='00FF00', name='Arial', strike=True) #strike:취소선
sheet['C1'].font= Font(color='0000FF', size=20, underline='single')

#4. 경계선(테두리) 그리기
from openpyxl.styles import Border,Side

border_thin = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
sheet['A1'].border = border_thin
sheet['B1'].border = border_thin
sheet['C1'].border = border_thin


#5. 모든 셀의 글씨를 가운데 정렬 및 80점 이상인 셀들은 노란색으로 표식!
from openpyxl.styles import Alignment, PatternFill

# 엑셀에서 각 기능능 분리가 되어 있음.

for row in sheet.rows:
    for cell in row:
        #가운데 정렬
        cell.alignment = Alignment(horizontal='center',vertical='center')

        if cell.column==1 : # 첫번째 칸(번호칸) 은 점수가 아니니 계산하지 않도록
            continue

        # 번호 제외하고.. 국어 영어 점수들 중 80점 이상인 셀의 배경색을 노랑색으로....
        if isinstance(cell.value, int) and cell.value>= 80: # 인스턴스가 int형이냐?
            cell.fill = PatternFill(fgColor='FFFF00', fill_type='solid')
            cell.font = Font(color = 'FF0000')

#데이터가 많아지면 스크롤 됨. .이때 특정 위치를 고정시키기 [틀 고정]
sheet.freeze_panes='B2' # B2를 기준으로 왼쪽, 위쪽이 그대로 고정됨. 
#-----------------------------------
wb.save('./01_excel/sample4_style.xlsx')
wb.close()
