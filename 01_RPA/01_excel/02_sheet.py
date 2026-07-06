# sheet 만드는 것 해 보자.

from openpyxl import Workbook
wb = Workbook()

#현재 활성화된 sheet 다음에 새로운 sheet 만들기
sheet = wb.create_sheet() # 새로운 sheet 만들어짐. 기본 이름은 'Sheet번호" 이다. 1번이 있으면 2번 ,


#sheet 의 제목 변경
sheet.title = 'My Sheet'
sheet.sheet_properties.tabColor = 'FF6600' # RGB 

#새로운 시트를 추가하면서 제목까지 지정
sheet2=wb.create_sheet('First Sheet')
sheet3=wb.create_sheet('Second Sheet',2) # 2번 인덱스(3번째) 위치에 새로운 시트 생성


# 특정 시트를 만들때.. 혹시 참조변수가 없었다면...
wb.create_sheet('Third')
sheet4 = wb['Third'] #인덱싱을 통해 참조 가능함.
sheet4['A1']='Hello'


#모든 sheet들의 이름들을 알수 있음.
print(wb.sheetnames)

#sheet를 복사하여 새로운 시트 만들기
new_sheet=wb.copy_worksheet(sheet4)
new_sheet.title = '복사된 시트'

#작업 완료 후 저장 -------------------------------------
wb.save('./01_excel/sample2.xlsx')
wb.close()