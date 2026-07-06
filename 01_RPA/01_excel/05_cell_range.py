from openpyxl import Workbook
wb = Workbook()
sheet = wb.active
#---------------------------------------

#1. 데이타를 한줄씩 입력하기 - 대량의 데이터를 리스트[]로 입력
sheet.append(['번호','국어','영어']) # 제목줄

#2. 데이터를 랜덤하게 만들어서 적용
import random
for no in range(1,11): #1~10
    sheet.append([no,random.randint(0,100),random.randint(0,100)]) # 성적은 목(더미) 데이타


#3. 영어성적(영어컬룸:C열) : A열 번호 , B열 국어 성적
col_eng = sheet['C']
for cell in col_eng:
    print(cell.value)
print()

#4. [국어,영어] 컬룸의 데이터를 가져오기 [ 엑셀의 슬라이스 명..]
col_list = sheet['B:C'] # B, C 열
for col in col_list: #하나의 컬룸씩 반복 
    for cell in col:
        print(cell.value, end = " ")
    print()
print()

# 5. 한줄 데이타 가져오기
# 시트의 첫줄에 제목글씨들이 있음.. 일단 이것 부터 가져와 보기
row_title = sheet[1] #1번 row 번호를 가진 한줄 가져오기
for cell in row_title:
    print(cell.value, end=" ")
print()

#6. 슬라이싱을 이용하여 여러줄의 데이터를 한줄씩 가져오기
row_list = sheet[2:6] # 엑셀의 슬라이싱은 마지믹 번호 포함 2행 ~6행

for row in row_list:
    for cell in row:
        print(cell.value, end=' ')
    print()
print()

#2번 줄 부터 끝줄까지..
row_list= sheet[2:sheet.max_row]
for row in row_list:
    for cell in row:
        print(cell.value, end=' ')
    print()
print()

#7. [2교시 끝 , 3교시 시작]
rows = sheet.rows #슬라이싱보다 편함 .. row들을 다 줌.
print(tuple(rows))
print()  

cols = sheet.columns
print( tuple(cols))
print('~'*30)
print()


# 이 튜플을 이용하여 '컬룸명, 인덱스명'을 사용하지 않고 값들 접근하기.
for row in tuple(sheet.rows):
    print(row[0].value)  # 번호칸
print()

for col in tuple(sheet.columns):
    print(col[0].value)  # 각 컬퓸묶음의 첫번째 요소 -- 제목줄 cell (번호,국어, 영어)
    print(col[1].value)  # 각 컬륨묶음의 두변찌 요소 -- 1번 줄 cell ( 1, xx, xx)


#위 튜플 방식 대신 사용하는 기능함수 iter_rows()
# for row in sheet.rows : # 이것은 tuple로 줌
for row in sheet.iter_rows():
    print(row[0].value) # 각 row의 첫번째 위치 값 [ 'A1', 'A2', 'A3....]

# iter_rows의 장점( sheet.rows와 다르게 시작과 종료위치를 지정할 수 있음. 무조건 다 가져옴. 제목줄까지..)
for row in sheet.iter_rows(min_row=2, max_row=6,min_col=2): # 2번줄부터 ~6번줄까지... 2번 줄부터
    print(row[0].value)


#---------------------------------------
wb.save('./01_excel/sample4.xlsx')
wb.close()
