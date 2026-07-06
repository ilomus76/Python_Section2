from openpyxl import load_workbook

wb = load_workbook('./01_excel/sample5.xlsx')
sheet = wb.active

#각 줄의 모든 셀값들을 읽어오기
for row in sheet.values: # 값들을 가져옴
    for cell in row:
        print(cell , end = " ")
    print()
print('-'*30)

# 결과치가 수식이 나옴.
# =SUM(10,20,30) 
# =AVERAGE(10,20,30) 
# 100 
# 200 
# 300 
# =sum(A3:A5) 

#셀에 수식이 있으면 수식자체가 읽어짐. 근데. 업무에 사용할때는 수식보다 계산된 값이 필요.
# 엑셀파일을 읽을때.. 데이타값으로만 읽어달라는 설정이 필요함. 
wb2 = load_workbook('./01_excel/sample5.xlsx',data_only=True)
sheet2=wb2.active
for row in sheet2.values:
    for cell in row:
        print(cell, end=' ')
    print()
print()

#엑셀파일을 프로그램을 만들고 읽으면 수식값이 None으로 읽어짐.
#사람이 한번은 그 엑셀파일을 직접 확인 한 후 저장해야 올바로 읽어짐